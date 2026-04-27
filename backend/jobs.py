"""In-memory job manager for the detect-confirm-recognize flow.

Each job lives in /tmp/tablex_jobs/<id>/ and holds:
  - page_001.png, page_002.png, ...   rasterized input pages
  - table_<n>.png                      perspective-warped table crops (after recognize)
  - table_<n>.csv                      per-table CSV (after recognize)

In-memory `_JOBS` dict tracks status + cached metadata. Garbage-collected on
TTL expiry by `cleanup_expired()` (called opportunistically on each request).
"""
from __future__ import annotations

import asyncio
import io
import os
import shutil
import threading
import time
import uuid
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

import cv2
import numpy as np
import pypdfium2 as pdfium
from PIL import Image

from pipeline import get_pipeline


JOB_ROOT = Path("/tmp/tablex_jobs")
JOB_TTL_SECONDS = 60 * 60  # 1 hour
PDF_DPI = 200


@dataclass
class TableResult:
    index: int
    page_index: int
    html: str
    csv: str
    cell_count: int


@dataclass
class Job:
    id: str
    created_at: float
    status: str = "detected"        # detected | running | done | error | cancelled
    progress: float = 0.0
    error: str | None = None
    pages: list[dict] = field(default_factory=list)   # detect output
    tables: list[TableResult] = field(default_factory=list)
    cancel_requested: bool = False


_JOBS: dict[str, Job] = {}
_LOCK = threading.Lock()
_STARTED_AT = time.time()

# Single-slot semaphore: inference (YOLO/TSR/OCR) is not safe to run in
# parallel on the shared singleton models, and the HF Space only has 2 vCPU
# anyway — concurrent inference just thrashes. All inference goes through
# `inference_slot()`. Endpoints stay async, so waiters suspend on the
# semaphore instead of holding a threadpool slot.
_inference_sem: asyncio.Semaphore | None = None


def inference_slot() -> asyncio.Semaphore:
    global _inference_sem
    if _inference_sem is None:
        _inference_sem = asyncio.Semaphore(1)
    return _inference_sem


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def _job_dir(job_id: str) -> Path:
    return JOB_ROOT / job_id


def _bbox_to_quad(bbox: list[int]) -> list[list[float]]:
    x1, y1, x2, y2 = bbox
    return [[x1, y1], [x2, y1], [x2, y2], [x1, y2]]


def cleanup_expired() -> None:
    now = time.time()
    expired: list[str] = []
    with _LOCK:
        for jid, job in _JOBS.items():
            if now - job.created_at > JOB_TTL_SECONDS:
                expired.append(jid)
        for jid in expired:
            _JOBS.pop(jid, None)
    for jid in expired:
        shutil.rmtree(_job_dir(jid), ignore_errors=True)


def get_job(job_id: str) -> Job | None:
    return _JOBS.get(job_id)


# ---------------------------------------------------------------------------
# Rasterization
# ---------------------------------------------------------------------------
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tif", ".tiff"}


def _rasterize_pdf(pdf_bytes: bytes, out_dir: Path, base: int) -> list[Path]:
    pdf = pdfium.PdfDocument(pdf_bytes)
    scale = PDF_DPI / 72.0
    paths: list[Path] = []
    for i, page in enumerate(pdf, start=1):
        pil = page.render(scale=scale).to_pil().convert("RGB")
        path = out_dir / f"page_{base + i:04d}.png"
        pil.save(path, format="PNG")
        paths.append(path)
        page.close()
    pdf.close()
    return paths


def _save_image(image_bytes: bytes, out_dir: Path, base: int) -> list[Path]:
    pil = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    path = out_dir / f"page_{base + 1:04d}.png"
    pil.save(path, format="PNG")
    return [path]


def _unpack_zip(zip_bytes: bytes, out_dir: Path, base: int) -> list[Path]:
    paths: list[Path] = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        members = sorted(
            (m for m in zf.namelist() if not m.endswith("/")),
            key=lambda s: s.lower(),
        )
        for name in members:
            ext = Path(name).suffix.lower()
            if ext not in IMAGE_EXTS:
                continue
            data = zf.read(name)
            try:
                pil = Image.open(io.BytesIO(data)).convert("RGB")
            except Exception:
                continue
            path = out_dir / f"page_{base + len(paths) + 1:04d}.png"
            pil.save(path, format="PNG")
            paths.append(path)
    return paths


def _ingest_one(
    data: bytes, content_type: str, out_dir: Path, base: int
) -> list[Path]:
    if content_type == "application/pdf":
        return _rasterize_pdf(data, out_dir, base)
    if content_type in {"application/zip", "application/x-zip-compressed"}:
        return _unpack_zip(data, out_dir, base)
    return _save_image(data, out_dir, base)


# ---------------------------------------------------------------------------
# create_job: rasterize + accumulate pages from one or more uploads
# ---------------------------------------------------------------------------
def create_job(uploads: list[tuple[bytes, str]]) -> Job:
    """`uploads` = list of (bytes, content_type) — concatenates pages in order."""
    cleanup_expired()
    job_id = uuid.uuid4().hex[:12]
    out_dir = _job_dir(job_id)
    out_dir.mkdir(parents=True, exist_ok=True)

    page_paths: list[Path] = []
    for data, content_type in uploads:
        added = _ingest_one(data, content_type, out_dir, base=len(page_paths))
        page_paths.extend(added)

    if not page_paths:
        raise ValueError("No usable pages or images in upload.")

    pages_meta: list[dict] = []
    for i, p in enumerate(page_paths):
        bgr = cv2.imread(str(p))
        h, w = bgr.shape[:2]
        pages_meta.append({
            "index": i,
            "filename": p.name,
            "width": w,
            "height": h,
            "detections": [],
            "detected": False,
        })

    job = Job(id=job_id, created_at=time.time(), pages=pages_meta)
    with _LOCK:
        _JOBS[job_id] = job
    return job


# ---------------------------------------------------------------------------
# add/remove pages on an existing job (mid-parse "Add more files" flow)
# ---------------------------------------------------------------------------
def _next_page_filename_index(job: "Job") -> int:
    used = 0
    for p in job.pages:
        try:
            n = int(p["filename"].split("_")[1].split(".")[0])
            used = max(used, n)
        except Exception:
            pass
    return used


def add_pages_to_job(job_id: str, uploads: list[tuple[bytes, str]]) -> Job:
    job = _JOBS.get(job_id)
    if job is None:
        raise ValueError("job not found")
    out_dir = _job_dir(job_id)
    out_dir.mkdir(parents=True, exist_ok=True)

    base = _next_page_filename_index(job)
    new_paths: list[Path] = []
    for data, content_type in uploads:
        added = _ingest_one(data, content_type, out_dir, base=base + len(new_paths))
        new_paths.extend(added)

    if not new_paths:
        raise ValueError("No usable pages in added files.")

    start = len(job.pages)
    for offset, p in enumerate(new_paths):
        bgr = cv2.imread(str(p))
        h, w = bgr.shape[:2]
        job.pages.append({
            "index": start + offset,
            "filename": p.name,
            "width": w,
            "height": h,
            "detections": [],
            "detected": False,
        })
    return job


def remove_page(job_id: str, page_index: int) -> Job:
    job = _JOBS.get(job_id)
    if job is None:
        raise ValueError("job not found")
    if page_index < 0 or page_index >= len(job.pages):
        raise ValueError("page index out of range")

    out_dir = _job_dir(job_id)

    keep_tables: list[TableResult] = []
    for t in job.tables:
        if t.page_index == page_index:
            for ext in (".png", ".csv"):
                f = out_dir / f"table_{t.index}{ext}"
                if f.exists():
                    try:
                        f.unlink()
                    except Exception:
                        pass
            continue
        if t.page_index > page_index:
            t.page_index -= 1
        keep_tables.append(t)
    job.tables = keep_tables

    removed = job.pages.pop(page_index)
    fpath = out_dir / removed["filename"]
    if fpath.exists():
        try:
            fpath.unlink()
        except Exception:
            pass
    for i, p in enumerate(job.pages):
        p["index"] = i
    return job


# ---------------------------------------------------------------------------
# detect: run YOLO detection on selected pages (None = all)
# ---------------------------------------------------------------------------
def run_detect(job_id: str, page_indices: list[int] | None) -> None:
    job = _JOBS.get(job_id)
    if job is None:
        return
    job.cancel_requested = False
    pipeline = get_pipeline()
    out_dir = _job_dir(job_id)
    indices = page_indices if page_indices is not None else list(range(len(job.pages)))
    for i in indices:
        if job.cancel_requested:
            break
        if i < 0 or i >= len(job.pages):
            continue
        page = job.pages[i]
        bgr = cv2.imread(str(out_dir / page["filename"]))
        detections = pipeline.detect_boxes(bgr)
        page["detections"] = [
            {"quad": _bbox_to_quad(d["bbox"]), "score": d["score"]}
            for d in detections
        ]
        page["detected"] = True
    job.cancel_requested = False


# ---------------------------------------------------------------------------
# recognize: warp + TSR + OCR per confirmed quad (background task)
# ---------------------------------------------------------------------------
def run_recognize(job_id: str, confirmed: list[dict]) -> None:
    """`confirmed` = [{page_index, quad: [[x,y]×4]}, ...]"""
    job = _JOBS.get(job_id)
    if job is None:
        return
    job.status = "running"
    job.progress = 0.0
    job.error = None
    job.cancel_requested = False
    start_index = len(job.tables)  # append, don't wipe — supports per-page parsing

    pipeline = get_pipeline()
    out_dir = _job_dir(job_id)
    total = max(1, len(confirmed))

    page_cache: dict[int, np.ndarray] = {}
    failures: list[str] = []
    cancelled = False
    for offset, item in enumerate(confirmed, start=1):
        if job.cancel_requested:
            cancelled = True
            break
        idx = start_index + offset
        try:
            page_index = int(item["page_index"])
            quad = item["quad"]
            if page_index not in page_cache:
                page_cache[page_index] = cv2.imread(
                    str(out_dir / job.pages[page_index]["filename"])
                )
            bgr = page_cache[page_index]
            r = pipeline.recognize_quad(bgr, quad)
            crop_path = out_dir / f"table_{idx}.png"
            csv_path = out_dir / f"table_{idx}.csv"
            cv2.imwrite(str(crop_path), r["crop_bgr"])
            csv_path.write_text(r["csv"], encoding="utf-8")
            job.tables.append(TableResult(
                index=idx,
                page_index=page_index,
                html=r["html"],
                csv=r["csv"],
                cell_count=len(r["cells"]),
            ))
        except Exception as e:
            failures.append(f"page {item.get('page_index')} table {idx}: {e}")
        job.progress = offset / total

    job.status = "cancelled" if cancelled else "done"
    job.progress = 1.0 if not cancelled else job.progress
    job.error = "; ".join(failures) if failures else None
    job.cancel_requested = False


# ---------------------------------------------------------------------------
# bundle: return zip bytes
# ---------------------------------------------------------------------------
def build_zip(job_id: str) -> bytes:
    job = _JOBS.get(job_id)
    if job is None or job.status != "done":
        raise ValueError("job not ready")
    out_dir = _job_dir(job_id)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for t in job.tables:
            csv_path = out_dir / f"table_{t.index}.csv"
            if csv_path.exists():
                zf.write(csv_path, arcname=f"table_{t.index}.csv")
    return buf.getvalue()


def request_cancel(job_id: str) -> bool:
    job = _JOBS.get(job_id)
    if job is None:
        return False
    job.cancel_requested = True
    return True


def build_combined_xlsx(job_id: str) -> bytes:
    """One workbook, one sheet per table. CSV-derived (backend/ doesn't track cells)."""
    from openpyxl import Workbook
    import csv as _csv

    job = _JOBS.get(job_id)
    if job is None:
        raise ValueError("job not found")
    if not job.tables:
        raise ValueError("no tables")

    wb = Workbook()
    wb.remove(wb.active)
    for i, t in enumerate(job.tables, start=1):
        ws = wb.create_sheet(title=f"Table {i}"[:31])
        reader = _csv.reader(io.StringIO(t.csv))
        for r_idx, row in enumerate(reader, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_combined_html(job_id: str) -> str:
    """Single styled HTML page rendering every table back-to-back."""
    job = _JOBS.get(job_id)
    if job is None:
        raise ValueError("job not found")
    if not job.tables:
        raise ValueError("no tables")

    sections: list[str] = []
    for i, t in enumerate(job.tables, start=1):
        sections.append(
            f'<section class="t"><h2>Table {i} '
            f'<span class="meta">page {t.page_index + 1}</span></h2>'
            f'{t.html}</section>'
        )
    body = "\n".join(sections)
    return (
        '<!doctype html><html lang="en"><head><meta charset="utf-8">'
        f'<title>Tables — job {job_id}</title>'
        '<style>'
        'body{font-family:-apple-system,Segoe UI,Roboto,sans-serif;'
        'background:#fafafa;color:#111;margin:0;padding:32px;}'
        '.t{background:#fff;border:1px solid #e5e7eb;border-radius:8px;'
        'padding:20px;margin-bottom:24px;box-shadow:0 1px 2px rgba(0,0,0,0.04);}'
        'h2{margin:0 0 12px;font-size:16px;color:#111;}'
        '.meta{color:#6b7280;font-weight:400;font-size:13px;margin-left:8px;}'
        'table{border-collapse:collapse;width:auto;}'
        'th,td{border:1px solid #d1d5db;padding:6px 10px;text-align:left;'
        'vertical-align:top;font-size:13px;}'
        'th{background:#f3f4f6;font-weight:600;}'
        '</style></head><body>'
        f'<h1>Extracted tables ({len(job.tables)})</h1>{body}'
        '</body></html>'
    )


def _cgroup_memory() -> tuple[int, int] | None:
    """Container memory quota (used, limit) from cgroup, or None."""
    try:
        with open("/sys/fs/cgroup/memory.current") as f:
            used = int(f.read().strip())
        with open("/sys/fs/cgroup/memory.max") as f:
            raw = f.read().strip()
            limit = int(raw) if raw != "max" else 0
        if limit > 0:
            return used, limit
    except Exception:
        pass
    try:
        with open("/sys/fs/cgroup/memory/memory.usage_in_bytes") as f:
            used = int(f.read().strip())
        with open("/sys/fs/cgroup/memory/memory.limit_in_bytes") as f:
            limit = int(f.read().strip())
        if 0 < limit < (1 << 62):
            return used, limit
    except Exception:
        pass
    return None


def get_health() -> dict:
    """In-container health: cpu/ram/uptime + active jobs.

    RAM prefers cgroup (real container quota) over psutil (host total).
    """
    info: dict = {
        "status": "ok",
        "uptime_s": int(time.time() - _STARTED_AT),
        "active_jobs": len(_JOBS),
    }
    try:
        import psutil
        info["cpu_percent"] = round(psutil.cpu_percent(interval=None), 1)
        info["process_rss_mb"] = int(psutil.Process().memory_info().rss / (1024 * 1024))
    except Exception as e:
        info["psutil_error"] = str(e)

    cg = _cgroup_memory()
    if cg is not None:
        used, limit = cg
        info["ram_used_mb"] = int(used / (1024 * 1024))
        info["ram_total_mb"] = int(limit / (1024 * 1024))
        info["ram_percent"] = round(used / limit * 100, 1)
        info["ram_source"] = "cgroup"
    else:
        try:
            import psutil
            vm = psutil.virtual_memory()
            info["ram_percent"] = round(vm.percent, 1)
            info["ram_used_mb"] = int(vm.used / (1024 * 1024))
            info["ram_total_mb"] = int(vm.total / (1024 * 1024))
            info["ram_source"] = "host"
        except Exception:
            pass
    return info


def build_combined_csv(job_id: str) -> str:
    job = _JOBS.get(job_id)
    if job is None:
        raise ValueError("job not found")
    if not job.tables:
        raise ValueError("no tables")
    parts: list[str] = []
    for i, t in enumerate(job.tables, start=1):
        parts.append(f"# Table {i} (page {t.page_index + 1})")
        parts.append(t.csv.rstrip())
        parts.append("")  # blank line separator
    return "\n".join(parts).rstrip() + "\n"


def build_page_csv_zip(job_id: str, page_index: int) -> bytes:
    job = _JOBS.get(job_id)
    if job is None:
        raise ValueError("job not found")
    out_dir = _job_dir(job_id)
    tables = [t for t in job.tables if t.page_index == page_index]
    if not tables:
        raise ValueError("no tables for page")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for t in tables:
            csv_path = out_dir / f"table_{t.index}.csv"
            if csv_path.exists():
                zf.write(csv_path, arcname=f"table_{t.index}.csv")
    return buf.getvalue()


def page_image_path(job_id: str, page_index: int) -> Path | None:
    job = _JOBS.get(job_id)
    if job is None or page_index < 0 or page_index >= len(job.pages):
        return None
    return _job_dir(job_id) / job.pages[page_index]["filename"]


def table_image_path(job_id: str, table_index: int) -> Path | None:
    if not _JOBS.get(job_id):
        return None
    p = _job_dir(job_id) / f"table_{table_index}.png"
    return p if p.exists() else None
